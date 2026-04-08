"""
FileIndex  ·  Enterprise Edition
──────────────────────────────────
核心差异化（对标投资委员会意见）：
  ① 文件内容全文检索  — Spotlight/Everything 只搜文件名，我们搜内容
  ② SQLite FTS5      — 毫秒级内容检索 + snippet 高亮摘要
  ③ 审计日志         — 企业合规必需，记录所有搜索和访问行为
  ④ 内容提取管线     — PDF / Word / Excel / 代码 / 纯文本，50+ 格式
  ⑤ 数据本地化       — 零云端依赖，数据不出本机（GGV 核心论点）
  ⑥ 所有 v2/v3 功能  — NLQ 解析、拼音、模糊、并行扫描、SQLite 持久化
"""

from flask import Flask, request, jsonify, send_file, Response, stream_with_context
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
import os, re, csv, time, platform, threading, queue
import json, uuid, sqlite3, fnmatch, io, hashlib, mimetypes
import getpass

# ── Optional deps ──────────────────────────────────────────────────────────────
try:
    from rapidfuzz import fuzz
    def _fuzzy(a, b): return fuzz.partial_ratio(a.lower(), b.lower()) / 100.0
    FUZZY_ENGINE = 'rapidfuzz'
except ImportError:
    from difflib import SequenceMatcher
    def _fuzzy(a, b): return SequenceMatcher(None, a.lower(), b.lower()).ratio()
    FUZZY_ENGINE = 'difflib'

try:
    from pypinyin import lazy_pinyin, Style
    def _pinyin(t): return ''.join(lazy_pinyin(t, style=Style.NORMAL))
    def _initials(t): return ''.join(lazy_pinyin(t, style=Style.FIRST_LETTER))
    PINYIN_OK = True
except ImportError:
    def _pinyin(t): return ''
    def _initials(t): return ''
    PINYIN_OK = False

# Content extraction deps
try:
    from pypdf import PdfReader
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    from docx import Document as DocxDoc
    DOCX_OK = True
except ImportError:
    DOCX_OK = False

try:
    from openpyxl import load_workbook
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ── App ────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DB_PATH       = os.path.join(BASE_DIR, 'fileindex.db')
SETTINGS_PATH = os.path.join(BASE_DIR, 'settings.json')

_db_lock   = threading.Lock()
_tasks     = {}
_task_lock = threading.Lock()
_cache     = {}
_clk       = threading.Lock()
CACHE_TTL  = 3600
CACHE_MAX  = 20

# ── Settings ───────────────────────────────────────────────────────────────────
DEFAULTS = {
    'ignorePatterns': ['node_modules','.git','.svn','__pycache__',
                       '.DS_Store','Thumbs.db','*.tmp'],
    'apiKey': '', 'fuzzyThreshold': 55,
}

def _cfg():
    if os.path.exists(SETTINGS_PATH):
        try:
            with open(SETTINGS_PATH,'r',encoding='utf-8') as f:
                return {**DEFAULTS, **json.load(f)}
        except Exception: pass
    return dict(DEFAULTS)

def _save_cfg(s):
    with open(SETTINGS_PATH,'w',encoding='utf-8') as f:
        json.dump(s, f, ensure_ascii=False, indent=2)

def _auth_ok():
    key = _cfg().get('apiKey','').strip()
    if not key: return True
    return request.headers.get('X-API-Key','') == key

# ── Database ───────────────────────────────────────────────────────────────────
def db():
    c = sqlite3.connect(DB_PATH, check_same_thread=False)
    c.row_factory = sqlite3.Row
    c.execute('PRAGMA journal_mode=WAL')
    c.execute('PRAGMA foreign_keys=ON')
    c.execute('PRAGMA synchronous=NORMAL')
    return c

def init_db():
    with _db_lock:
        c = db()
        c.executescript('''
            CREATE TABLE IF NOT EXISTS scan_indexes(
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                name            TEXT NOT NULL,
                directory       TEXT NOT NULL,
                created_at      TEXT NOT NULL,
                total_files     INTEGER DEFAULT 0,
                skipped_files   INTEGER DEFAULT 0,
                content_files   INTEGER DEFAULT 0,
                status          TEXT DEFAULT "scanning",
                include_content INTEGER DEFAULT 0,
                ignore_patterns TEXT DEFAULT "[]"
            );
            CREATE TABLE IF NOT EXISTS files(
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                index_id         INTEGER NOT NULL,
                path             TEXT NOT NULL,
                name             TEXT NOT NULL,
                ext              TEXT,
                size_kb          REAL,
                modified_time    TEXT,
                modified_ts      INTEGER,
                creator          TEXT,
                name_pinyin      TEXT,
                pinyin_initials  TEXT,
                size_hash        TEXT,
                has_content      INTEGER DEFAULT 0,
                FOREIGN KEY(index_id) REFERENCES scan_indexes(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_fi    ON files(index_id);
            CREATE INDEX IF NOT EXISTS idx_ext   ON files(index_id,ext);
            CREATE INDEX IF NOT EXISTS idx_sz    ON files(index_id,size_kb);
            CREATE INDEX IF NOT EXISTS idx_mt    ON files(index_id,modified_ts);
            CREATE INDEX IF NOT EXISTS idx_dup   ON files(index_id,size_hash);

            CREATE TABLE IF NOT EXISTS file_contents(
                file_id  INTEGER PRIMARY KEY,
                content  TEXT,
                FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
            );
            CREATE VIRTUAL TABLE IF NOT EXISTS content_fts USING fts5(
                content,
                content="file_contents",
                content_rowid="file_id",
                tokenize="unicode61"
            );
            CREATE TRIGGER IF NOT EXISTS fc_ai AFTER INSERT ON file_contents BEGIN
                INSERT INTO content_fts(rowid, content) VALUES (new.file_id, new.content);
            END;
            CREATE TRIGGER IF NOT EXISTS fc_ad AFTER DELETE ON file_contents BEGIN
                INSERT INTO content_fts(content_fts, rowid, content)
                VALUES ("delete", old.file_id, old.content);
            END;

            CREATE TABLE IF NOT EXISTS audit_log(
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                index_id     INTEGER,
                action       TEXT NOT NULL,
                query        TEXT,
                mode         TEXT,
                result_count INTEGER,
                duration_ms  INTEGER,
                timestamp    TEXT NOT NULL,
                ip           TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_al_idx ON audit_log(index_id);
            CREATE INDEX IF NOT EXISTS idx_al_ts  ON audit_log(timestamp);
        ''')
        c.commit(); c.close()

init_db()

# ── Cache ──────────────────────────────────────────────────────────────────────
def _cget(iid):
    with _clk:
        e = _cache.get(iid)
        if e and time.time()-e['ts'] < CACHE_TTL: return e['d']
        if e: del _cache[iid]
        return None

def _cset(iid, d):
    with _clk:
        if len(_cache) >= CACHE_MAX and iid not in _cache:
            del _cache[min(_cache, key=lambda k: _cache[k]['ts'])]
        _cache[iid] = {'d': d, 'ts': time.time()}

def _cdel(iid):
    with _clk: _cache.pop(iid, None)

def _audit(index_id, action, query='', mode='', result_count=0, duration_ms=0):
    """Record every search/access to audit_log."""
    try:
        with _db_lock:
            c = db()
            c.execute(
                'INSERT INTO audit_log(index_id,action,query,mode,result_count,duration_ms,timestamp,ip) '
                'VALUES(?,?,?,?,?,?,?,?)',
                (index_id, action, query, mode, result_count, duration_ms,
                 time.strftime('%Y-%m-%d %H:%M:%S'), request.remote_addr or '')
            )
            c.commit(); c.close()
    except Exception: pass

# ── Content extraction ─────────────────────────────────────────────────────────
TEXT_EXTS = {
    '.txt','.md','.py','.js','.ts','.jsx','.tsx','.mjs','.cjs',
    '.css','.scss','.less','.html','.htm','.xml','.svg',
    '.json','.jsonl','.yaml','.yml','.toml','.ini','.cfg','.conf','.env',
    '.sh','.bash','.zsh','.fish','.bat','.ps1',
    '.go','.rs','.java','.kt','.swift','.cpp','.c','.h','.hpp',
    '.rb','.php','.lua','.r','.m','.scala','.cs','.vb',
    '.sql','.graphql','.proto',
    '.csv','.tsv','.log','.gitignore','.dockerfile',
    '.tex','.rst','.org',
}
CONTENT_MAX_KB = 2048   # skip files > 2 MB for content extraction

def extract_text(path: str, ext: str, size_kb: float) -> str | None:
    """Extract plain text from a file. Returns None if not extractable."""
    if size_kb > CONTENT_MAX_KB:
        return None

    # Plain text / code
    if ext in TEXT_EXTS:
        for enc in ('utf-8', 'gb18030', 'latin-1'):
            try:
                with open(path, 'r', encoding=enc, errors='strict') as f:
                    content = f.read(200_000)   # cap at 200K chars
                return content.strip() or None
            except (UnicodeDecodeError, UnicodeError):
                continue
        return None

    # PDF
    if ext == '.pdf' and PDF_OK:
        try:
            reader = PdfReader(path)
            parts = []
            for page in reader.pages[:50]:   # first 50 pages
                text = page.extract_text() or ''
                parts.append(text)
            content = '\n'.join(parts).strip()
            return content[:200_000] or None
        except Exception:
            return None

    # Word (.docx)
    if ext == '.docx' and DOCX_OK:
        try:
            doc = DocxDoc(path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            content = '\n'.join(paragraphs).strip()
            return content[:200_000] or None
        except Exception:
            return None

    # Excel (.xlsx)
    if ext in ('.xlsx', '.xlsm') and XLSX_OK:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets[:5]:      # first 5 sheets
                rows_text = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i > 500: break            # first 500 rows per sheet
                    cells = [str(c) for c in row if c is not None and str(c).strip()]
                    if cells:
                        rows_text.append(' | '.join(cells))
                parts.append(f'[{sheet.title}]\n' + '\n'.join(rows_text))
            content = '\n\n'.join(parts).strip()
            return content[:200_000] or None
        except Exception:
            return None

    return None

# ── NLQ Parser ─────────────────────────────────────────────────────────────────
_UNITS = {'kb':1,'k':1,'mb':1024,'m':1024,'gb':1024**2,'g':1024**2,'tb':1024**3}
_EXT_MAP = {
    'pdf文件':'.pdf','pdf':'.pdf','文档':'.docx','word文档':'.docx','word':'.docx',
    'excel表格':'.xlsx','excel':'.xlsx','表格':'.xlsx',
    'ppt':'.pptx','演示文稿':'.pptx','幻灯片':'.pptx',
    '图片':'.jpg','照片':'.jpg','png':'.png',
    '视频':'.mp4','视频文件':'.mp4','音频':'.mp3','音乐':'.mp3',
    '压缩包':'.zip','压缩文件':'.zip',
    'python文件':'.py','python代码':'.py',
    'js文件':'.js','ts文件':'.ts',
    '文本':'.txt','纯文本':'.txt','日志':'.log','日志文件':'.log',
    'markdown':'.md','md文件':'.md',
    'pdfs':'.pdf','pdf files':'.pdf',
    'word files':'.docx','word documents':'.docx',
    'excel files':'.xlsx','spreadsheets':'.xlsx',
    'powerpoint':'.pptx','slides':'.pptx','presentations':'.pptx',
    'images':'.jpg','photos':'.jpg','pictures':'.jpg',
    'videos':'.mp4','movies':'.mp4',
    'audio':'.mp3','music':'.mp3',
    'zip files':'.zip','archives':'.zip',
    'python files':'.py','scripts':'.py',
    'javascript files':'.js','js files':'.js',
    'typescript files':'.ts','text files':'.txt',
    'log files':'.log','logs':'.log','markdown files':'.md',
}

def parse_nlq(text: str) -> dict:
    q   = text.strip()
    now = datetime.now()
    out = {'keyword':'','ext':None,'min_size_kb':None,'max_size_kb':None,
           'date_after':None,'date_before':None,'tokens':[]}
    rem = q

    def unit(u): return _UNITS.get(u.lower(), 1)

    for pat, kind in [
        (r'(?:大于|超过|≥|>=?)\s*(\d+(?:\.\d+)?)\s*(TB|GB|MB|KB|[TGMK])\b', 'min'),
        (r'(\d+(?:\.\d+)?)\s*(TB|GB|MB|KB|[TGMK])\s*(?:以上|以上的)', 'min'),
        (r'(?:larger?\s+than|bigger?\s+than|more\s+than|over|above)\s+(\d+(?:\.\d+)?)\s*(TB|GB|MB|KB|[TGMK])\b', 'min'),
        (r'(?:小于|低于|≤|<=?)\s*(\d+(?:\.\d+)?)\s*(TB|GB|MB|KB|[TGMK])\b', 'max'),
        (r'(\d+(?:\.\d+)?)\s*(TB|GB|MB|KB|[TGMK])\s*(?:以下|以下的)', 'max'),
        (r'(?:smaller?\s+than|less\s+than|under|below)\s+(\d+(?:\.\d+)?)\s*(TB|GB|MB|KB|[TGMK])\b', 'max'),
    ]:
        m = re.search(pat, rem, re.IGNORECASE)
        if m:
            val = float(m.group(1)) * unit(m.group(2))
            if kind == 'min' and not out['min_size_kb']:
                out['min_size_kb'] = val
                out['tokens'].append({'type':'size','label':f'≥{m.group(1)}{m.group(2).upper()}'})
                rem = rem[:m.start()] + rem[m.end():]
            elif kind == 'max' and not out['max_size_kb']:
                out['max_size_kb'] = val
                out['tokens'].append({'type':'size','label':f'≤{m.group(1)}{m.group(2).upper()}'})
                rem = rem[:m.start()] + rem[m.end():]

    if re.search(r'\b(?:大文件|large\s*files?|huge\s*files?)\b', rem, re.I) and not out['min_size_kb']:
        out['min_size_kb'] = 50*1024
        out['tokens'].append({'type':'size','label':'≥50MB'})
        rem = re.sub(r'\b(?:大文件|large\s*files?|huge\s*files?)\b','',rem,flags=re.I)

    date_rules = [
        (r'\b(?:今天|今日|today)\b', now.replace(hour=0,minute=0,second=0), None),
        (r'\b(?:昨天|yesterday)\b', (now-timedelta(days=1)).replace(hour=0,minute=0,second=0), now.replace(hour=0,minute=0,second=0)),
        (r'\b(?:本周|this\s+week)\b', now-timedelta(days=now.weekday()), None),
        (r'\b(?:上周|last\s+week)\b', now-timedelta(days=now.weekday()+7), now-timedelta(days=now.weekday())),
        (r'\b(?:本月|this\s+month)\b', now.replace(day=1), None),
        (r'\b(?:上个月|last\s+month)\b', (now.replace(day=1)-timedelta(days=1)).replace(day=1), now.replace(day=1)),
        (r'\b(?:今年|this\s+year)\b', now.replace(month=1,day=1), None),
        (r'\b(?:最近一周|past\s+week)\b', now-timedelta(days=7), None),
        (r'\b(?:最近一个月|past\s+month)\b', now-timedelta(days=30), None),
    ]
    for pat, da, db in date_rules:
        if re.search(pat, rem, re.I):
            if da: out['date_after']  = da.strftime('%Y-%m-%d')
            if db: out['date_before'] = db.strftime('%Y-%m-%d')
            label = re.search(pat,rem,re.I).group(0).strip()
            out['tokens'].append({'type':'date','label':label})
            rem = re.sub(pat,'',rem,flags=re.I); break

    ym = re.search(r'\b(\d{4})年?\b', rem)
    if ym and not out['date_after']:
        yr = int(ym.group(1))
        out['date_after'] = f'{yr}-01-01'; out['date_before'] = f'{yr+1}-01-01'
        out['tokens'].append({'type':'date','label':f'{yr}年'})
        rem = rem[:ym.start()]+rem[ym.end():]

    em = re.search(r'\b(\.\w+)\b', rem)
    if em and not out['ext']:
        out['ext'] = em.group(1).lower()
        out['tokens'].append({'type':'ext','label':em.group(1)})
        rem = rem[:em.start()]+rem[em.end():]

    if not out['ext']:
        for kw in sorted(_EXT_MAP, key=len, reverse=True):
            if kw in rem.lower():
                out['ext'] = _EXT_MAP[kw]
                out['tokens'].append({'type':'ext','label':kw})
                rem = re.sub(re.escape(kw),'',rem,flags=re.I); break

    out['keyword'] = re.sub(r'\s+',' ',rem).strip().strip('的了和|')
    return out

# ── OS helpers ─────────────────────────────────────────────────────────────────
_OS = platform.system().lower()

def _owner(path, stat):
    try:
        if _OS == 'windows':
            try:
                import win32security as ws
                sd  = ws.GetFileSecurity(path, ws.OWNER_SECURITY_INFORMATION)
                sid = sd.GetSecurityDescriptorOwner()
                n,d,_ = ws.LookupAccountSid(None, sid); return f'{d}\\{n}'
            except ImportError: return getpass.getuser()
        else:
            try:
                import pwd; return pwd.getpwuid(stat.st_uid).pw_name
            except Exception: return getpass.getuser()
    except Exception: return getpass.getuser()

def _vpath(p):
    n = os.path.normpath(p)
    if not os.path.exists(n): raise ValueError('目录不存在')
    if not os.path.isdir(n):  raise ValueError('路径不是目录')
    return n

def _ignore(name, path, pats):
    return any(fnmatch.fnmatch(name,p) or fnmatch.fnmatch(path,p) for p in pats)

# ── Index worker ───────────────────────────────────────────────────────────────
def _index_worker(dir_path, task_id, iid, pats, include_content):
    with _task_lock:
        q = _tasks.get(task_id)
    if not q: return

    def push(t, **kw): q.put({'type':t, **kw})
    push('progress', message='开始扫描…', processed=0, skipped=0, phase='scan')

    batch, done, skip = [], 0, 0
    seen = set()
    dq   = queue.SimpleQueue()
    dq.put(dir_path)
    pool = ThreadPoolExecutor(max_workers=min(8,(os.cpu_count() or 2)*2))

    def _proc(e):
        try:
            if not os.access(e.path, os.R_OK): return None, True
            st = e.stat(follow_symlinks=False)
            if e.is_symlink(): return None, False
            if e.is_file(follow_symlinks=False):
                ext   = os.path.splitext(e.name)[1].lower() or '(无扩展名)'
                mts   = int(st.st_mtime)
                shash = hashlib.md5(f'{round(st.st_size/1024,1)}:{e.name}'.encode()).hexdigest()[:12]
                return {
                    'path'          : e.path,
                    'name'          : e.name,
                    'ext'           : ext,
                    'size_kb'       : round(st.st_size/1024, 2),
                    'modified_time' : time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mts)),
                    'modified_ts'   : mts,
                    'creator'       : _owner(e.path, st),
                    'name_pinyin'   : _pinyin(e.name) if PINYIN_OK else '',
                    'pinyin_initials': _initials(e.name) if PINYIN_OK else '',
                    'size_hash'     : shash,
                    'has_content'   : 0,
                }, False
        except (PermissionError, OSError): return None, True
        return None, False

    try:
        while not dq.empty():
            cur  = dq.get()
            real = os.path.realpath(cur)
            if real in seen: continue
            seen.add(real)
            try: entries = list(os.scandir(cur))
            except (PermissionError, OSError): skip += 1; continue

            dirs_, files_ = [], []
            for e in entries:
                if _ignore(e.name, e.path, pats): skip += 1; continue
                try:
                    if e.is_dir(follow_symlinks=False): dirs_.append(e)
                    elif e.is_file(follow_symlinks=False): files_.append(e)
                except OSError: skip += 1

            for d in dirs_: dq.put(d.path)
            futs = {pool.submit(_proc, f): f for f in files_}
            for fut in as_completed(futs):
                row, was_skip = fut.result()
                if was_skip: skip += 1
                elif row: batch.append(row); done += 1

            if len(batch) >= 2000:
                _flush_meta(iid, batch); batch.clear()
                push('progress', message=f'已扫描 {done:,} 个文件…',
                     processed=done, skipped=skip, phase='scan')
    finally:
        pool.shutdown(wait=False)

    if batch: _flush_meta(iid, batch)

    # ── Phase 2: content extraction ────────────────────────────────────────────
    content_done = 0
    if include_content:
        push('progress', message='开始提取文件内容…', processed=done, skipped=skip, phase='content')
        c    = db()
        rows = c.execute(
            'SELECT id, path, ext, size_kb FROM files WHERE index_id=? ORDER BY size_kb ASC',
            (iid,)
        ).fetchall()
        c.close()

        content_pool = ThreadPoolExecutor(max_workers=min(4, (os.cpu_count() or 2)))

        def _extract_one(row):
            text = extract_text(row['path'], row['ext'], row['size_kb'])
            if text:
                return row['id'], text
            return None

        cbatch = []
        for fut in as_completed({content_pool.submit(_extract_one, r): r for r in rows}):
            result = fut.result()
            if result:
                cbatch.append(result)
                content_done += 1
            if len(cbatch) >= 500:
                _flush_content(cbatch); cbatch.clear()
                push('progress',
                     message=f'已提取内容 {content_done:,} 个文件…',
                     processed=done, contentDone=content_done, skipped=skip, phase='content')
        if cbatch: _flush_content(cbatch)
        content_pool.shutdown(wait=False)

    # ── Finalize ───────────────────────────────────────────────────────────────
    with _db_lock:
        c = db()
        c.execute(
            'UPDATE scan_indexes SET status=?,total_files=?,skipped_files=?,content_files=? WHERE id=?',
            ('done', done, skip, content_done, iid)
        )
        c.commit(); c.close()

    push('done', indexId=iid, totalFiles=done, skippedFiles=skip,
         contentFiles=content_done,
         message=f'索引完成！{done:,} 个文件' +
                 (f'，已提取 {content_done:,} 份内容' if include_content else ''))

    def _clean():
        time.sleep(120)
        with _task_lock: _tasks.pop(task_id, None)
    threading.Thread(target=_clean, daemon=True).start()


def _flush_meta(iid, rows):
    with _db_lock:
        c = db()
        c.executemany(
            '''INSERT INTO files
               (index_id,path,name,ext,size_kb,modified_time,modified_ts,
                creator,name_pinyin,pinyin_initials,size_hash,has_content)
               VALUES(:index_id,:path,:name,:ext,:size_kb,:modified_time,:modified_ts,
                      :creator,:name_pinyin,:pinyin_initials,:size_hash,:has_content)''',
            [{**r, 'index_id': iid} for r in rows]
        )
        c.commit(); c.close()


def _flush_content(pairs):
    """pairs = [(file_id, content_text), ...]"""
    with _db_lock:
        c = db()
        c.executemany(
            'INSERT OR REPLACE INTO file_contents(file_id, content) VALUES(?,?)', pairs
        )
        c.execute(
            f'UPDATE files SET has_content=1 WHERE id IN ({",".join("?" for _ in pairs)})',
            [p[0] for p in pairs]
        )
        c.commit(); c.close()

# ── In-memory load & cache ─────────────────────────────────────────────────────
def _load(iid):
    cached = _cget(iid)
    if cached is not None: return cached
    c = db()
    rows = c.execute(
        '''SELECT path,name,ext,size_kb AS size,modified_time AS modifiedTime,
                  modified_ts,creator,name_pinyin,pinyin_initials,size_hash,has_content
           FROM files WHERE index_id=?''', (iid,)
    ).fetchall()
    c.close()
    d = [dict(r) for r in rows]
    _cset(iid, d); return d

# ── Auth guard ─────────────────────────────────────────────────────────────────
@app.before_request
def _guard():
    if request.method == 'OPTIONS': return
    if not _auth_ok(): return jsonify({'error':'未授权'}), 401

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route('/api/index', methods=['POST'])
def api_index():
    try:
        data            = request.get_json() or {}
        dir_path        = _vpath(data.get('directoryPath',''))
        name            = data.get('name','').strip() or os.path.basename(dir_path) or dir_path
        include_content = bool(data.get('includeContent', False))
        cfg             = _cfg()
        pats            = list(set(cfg.get('ignorePatterns',[]) + data.get('ignorePatterns',[])))

        with _db_lock:
            c   = db()
            cur = c.execute(
                'INSERT INTO scan_indexes(name,directory,created_at,status,include_content,ignore_patterns) '
                'VALUES(?,?,?,?,?,?)',
                (name, dir_path, time.strftime('%Y-%m-%d %H:%M:%S'),
                 'scanning', int(include_content), json.dumps(pats))
            )
            iid = cur.lastrowid; c.commit(); c.close()

        tid = uuid.uuid4().hex
        q   = queue.Queue()
        with _task_lock: _tasks[tid] = q
        threading.Thread(target=_index_worker, args=(dir_path,tid,iid,pats,include_content),
                         daemon=True).start()
        return jsonify({'success':True,'taskId':tid,'indexId':iid})
    except ValueError as e: return jsonify({'error':str(e)}), 400
    except Exception  as e: return jsonify({'error':str(e)}), 500


@app.route('/api/index/stream')
def api_stream():
    tid = request.args.get('taskId','')
    if not tid: return jsonify({'error':'缺少 taskId'}), 400
    with _task_lock: q = _tasks.get(tid)
    if q is None: return jsonify({'error':'任务不存在'}), 404
    def gen():
        while True:
            try:
                ev = q.get(timeout=30)
                yield f'data: {json.dumps(ev, ensure_ascii=False)}\n\n'
                if ev.get('type') in ('done','error'): break
            except queue.Empty: yield 'data: {"type":"ping"}\n\n'
    return Response(stream_with_context(gen()), mimetype='text/event-stream',
                    headers={'Cache-Control':'no-cache','X-Accel-Buffering':'no'})


@app.route('/api/indexes')
def api_list():
    c    = db()
    rows = c.execute(
        'SELECT id,name,directory,created_at,total_files,skipped_files,content_files,'
        'status,include_content FROM scan_indexes ORDER BY created_at DESC'
    ).fetchall()
    c.close()
    return jsonify({'success':True,'indexes':[dict(r) for r in rows]})


@app.route('/api/indexes/<int:iid>', methods=['DELETE'])
def api_del_index(iid):
    with _db_lock:
        c = db(); c.execute('DELETE FROM scan_indexes WHERE id=?',(iid,)); c.commit(); c.close()
    _cdel(iid)
    return jsonify({'success':True})


@app.route('/api/indexes/<int:iid>/rename', methods=['POST'])
def api_rename(iid):
    name = (request.get_json() or {}).get('name','').strip()
    if not name: return jsonify({'error':'名称不能为空'}), 400
    with _db_lock:
        c = db(); c.execute('UPDATE scan_indexes SET name=? WHERE id=?',(name,iid)); c.commit(); c.close()
    return jsonify({'success':True})


@app.route('/api/search', methods=['POST'])
def api_search():
    t0 = time.time()
    try:
        data       = request.get_json() or {}
        iid        = int(data.get('indexId',0))
        raw        = data.get('searchTerm','').strip()
        nlq_mode   = bool(data.get('nlq', False))
        ext_filter = data.get('extFilter','').lower().strip()
        sort_by    = data.get('sortBy','')
        sort_order = data.get('sortOrder','asc')
        page       = max(1, int(data.get('page',1)))
        pgsz       = min(500, max(1, int(data.get('pageSize',100))))
        fuzzy      = bool(data.get('fuzzy',False))

        if not iid: return jsonify({'error':'请提供 indexId'}), 400

        parsed     = parse_nlq(raw) if nlq_mode else None
        keyword    = (parsed['keyword'] if parsed else raw).lower()
        ext_f      = ext_filter or (parsed['ext'] if parsed else None)
        min_sz     = parsed['min_size_kb'] if parsed else None
        max_sz     = parsed['max_size_kb'] if parsed else None
        date_after = parsed['date_after']  if parsed else None
        date_before= parsed['date_before'] if parsed else None
        tokens     = parsed['tokens']      if parsed else []

        cfg       = _cfg()
        threshold = cfg.get('fuzzyThreshold',55) / 100.0
        all_files = _load(iid)

        def _match(row):
            if ext_f and row.get('ext','').lower() != ext_f: return False, 0.0
            if min_sz is not None and float(row.get('size') or 0) < min_sz: return False, 0.0
            if max_sz is not None and float(row.get('size') or 0) > max_sz: return False, 0.0
            if date_after  and (row.get('modifiedTime','') or '') < date_after:  return False, 0.0
            if date_before and (row.get('modifiedTime','') or '') >= date_before: return False, 0.0
            if not keyword: return True, 1.0
            nm  = (row.get('name','')    or '').lower()
            pt  = (row.get('path','')    or '').lower()
            ow  = (row.get('creator','') or '').lower()
            py  = (row.get('name_pinyin','')    or '').lower()
            pyi = (row.get('pinyin_initials','') or '').lower()
            if keyword in nm or keyword in pt or keyword in ow or keyword in py or keyword in pyi:
                return True, 1.0
            if fuzzy:
                sc = max(_fuzzy(keyword,nm), _fuzzy(keyword,py), _fuzzy(keyword,pyi))
                if sc >= threshold: return True, sc
            return False, 0.0

        scored = [(sc,r) for r in all_files for ok,sc in [_match(r)] if ok]
        total  = len(scored)

        if sort_by in ('name','size','modifiedTime','creator','ext'):
            rev = sort_order == 'desc'
            scored.sort(key=lambda x: float(x[1].get('size') or 0) if sort_by=='size'
                        else (x[1].get(sort_by) or ''), reverse=rev)
        elif fuzzy and keyword:
            scored.sort(key=lambda x: -x[0])

        page_rows   = scored[(page-1)*pgsz : page*pgsz]
        results     = [{'score':round(s*100), **r} for s,r in page_rows]
        total_pages = max(1,(total+pgsz-1)//pgsz)

        ext_cnt = {}
        for _,r in scored:
            e = r.get('ext') or '(无扩展名)'
            ext_cnt[e] = ext_cnt.get(e,0)+1
        ext_stats = sorted(ext_cnt.items(), key=lambda x:-x[1])[:15]

        dur = int((time.time()-t0)*1000)
        _audit(iid, 'search', raw, 'metadata', total, dur)

        return jsonify({
            'success':True,'results':results,'total':total,'page':page,
            'pageSize':pgsz,'totalPages':total_pages,'extStats':ext_stats,
            'tokens':tokens,'parsedFilters':{
                'keyword':keyword,'ext':ext_f,'minSizeKB':min_sz,'maxSizeKB':max_sz,
                'dateAfter':date_after,'dateBefore':date_before
            },
            'fuzzyEngine': FUZZY_ENGINE if fuzzy else None,
            'pinyinAvailable': PINYIN_OK,
            'durationMs': dur,
        })
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/search/content', methods=['POST'])
def api_search_content():
    """Full-text search inside file contents using SQLite FTS5 + snippet highlighting."""
    t0 = time.time()
    try:
        data   = request.get_json() or {}
        iid    = int(data.get('indexId',0))
        query  = data.get('query','').strip()
        page   = max(1, int(data.get('page',1)))
        pgsz   = min(200, max(1, int(data.get('pageSize',50))))
        ext_f  = data.get('extFilter','').lower().strip()

        if not iid:   return jsonify({'error':'请提供 indexId'}), 400
        if not query: return jsonify({'error':'请提供搜索内容'}), 400

        # Build FTS5 query (quote for safety)
        fts_query = ' '.join(f'"{w}"' if ' ' not in w else w for w in query.split())

        c = db()

        # Check content availability
        content_count = c.execute(
            'SELECT COUNT(*) FROM file_contents fc JOIN files f ON f.id=fc.file_id WHERE f.index_id=?',
            (iid,)
        ).fetchone()[0]

        if content_count == 0:
            c.close()
            return jsonify({'error':'该索引未提取文件内容，请重新建立索引并勾选「提取文件内容」','contentCount':0}), 400

        ext_clause = 'AND f.ext=?' if ext_f else ''
        params     = [iid, fts_query] + ([ext_f] if ext_f else [])

        sql = f'''
            SELECT
                f.id, f.path, f.name, f.ext,
                f.size_kb AS size,
                f.modified_time AS modifiedTime,
                f.creator,
                snippet(content_fts, 0, "<mark>", "</mark>", "…", 48) AS snippet,
                rank
            FROM content_fts
            JOIN file_contents fc ON fc.file_id = content_fts.rowid
            JOIN files f ON f.id = fc.file_id
            WHERE f.index_id=? AND content_fts MATCH ?
            {ext_clause}
            ORDER BY rank
            LIMIT {pgsz} OFFSET {(page-1)*pgsz}
        '''
        rows = c.execute(sql, params).fetchall()

        count_sql = f'''
            SELECT COUNT(*)
            FROM content_fts
            JOIN file_contents fc ON fc.file_id = content_fts.rowid
            JOIN files f ON f.id = fc.file_id
            WHERE f.index_id=? AND content_fts MATCH ?
            {ext_clause}
        '''
        total = c.execute(count_sql, params).fetchone()[0]
        c.close()

        results = [dict(r) for r in rows]
        dur     = int((time.time()-t0)*1000)
        _audit(iid, 'content_search', query, 'content', total, dur)

        return jsonify({
            'success'       : True,
            'results'       : results,
            'total'         : total,
            'page'          : page,
            'pageSize'      : pgsz,
            'totalPages'    : max(1,(total+pgsz-1)//pgsz),
            'contentCount'  : content_count,
            'durationMs'    : dur,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/<int:iid>')
def api_stats(iid):
    try:
        c    = db()
        meta = c.execute('SELECT * FROM scan_indexes WHERE id=?',(iid,)).fetchone()
        if not meta: c.close(); return jsonify({'error':'索引不存在'}), 404
        total_sz   = c.execute('SELECT COALESCE(SUM(size_kb),0) FROM files WHERE index_id=?',(iid,)).fetchone()[0]
        top_ext    = c.execute('SELECT ext,COUNT(*) cnt FROM files WHERE index_id=? GROUP BY ext ORDER BY cnt DESC LIMIT 12',(iid,)).fetchall()
        top_cr     = c.execute('SELECT creator,COUNT(*) cnt FROM files WHERE index_id=? GROUP BY creator ORDER BY cnt DESC LIMIT 8',(iid,)).fetchall()
        largest    = c.execute('SELECT path,name,size_kb FROM files WHERE index_id=? ORDER BY size_kb DESC LIMIT 5',(iid,)).fetchall()
        since_ts   = int((datetime.now()-timedelta(days=30)).timestamp())
        timeline   = c.execute(
            'SELECT strftime("%Y-%m-%d",modified_time) AS day,COUNT(*) cnt FROM files '
            'WHERE index_id=? AND modified_ts>=? GROUP BY day ORDER BY day', (iid,since_ts)
        ).fetchall()
        # Search frequency stats
        top_queries = c.execute(
            'SELECT query,COUNT(*) cnt FROM audit_log WHERE index_id=? AND action="search" '
            'AND query!="" GROUP BY query ORDER BY cnt DESC LIMIT 8', (iid,)
        ).fetchall()
        c.close()
        return jsonify({
            'success'       : True,
            'totalFiles'    : meta['total_files'],
            'totalSizeKB'   : round(total_sz,2),
            'contentFiles'  : meta['content_files'] or 0,
            'includeContent': bool(meta['include_content']),
            'topExtensions' : [[r['ext'],r['cnt']] for r in top_ext],
            'topCreators'   : [[r['creator'],r['cnt']] for r in top_cr],
            'largestFiles'  : [{'path':r['path'],'name':r['name'],'size':r['size_kb']} for r in largest],
            'timeline'      : [[r['day'],r['cnt']] for r in timeline],
            'topQueries'    : [[r['query'],r['cnt']] for r in top_queries],
        })
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/audit/<int:iid>')
def api_audit(iid):
    """Return audit log for an index. Enterprise compliance feature."""
    try:
        limit  = min(500, max(1, int(request.args.get('limit',100))))
        action = request.args.get('action','')
        c      = db()
        where  = 'WHERE index_id=?'
        params = [iid]
        if action: where += ' AND action=?'; params.append(action)
        rows = c.execute(
            f'SELECT id,action,query,mode,result_count,duration_ms,timestamp,ip '
            f'FROM audit_log {where} ORDER BY timestamp DESC LIMIT {limit}',
            params
        ).fetchall()
        total = c.execute(f'SELECT COUNT(*) FROM audit_log {where}', params).fetchone()[0]
        c.close()
        _audit(iid, 'audit_view', '', 'audit', 0, 0)
        return jsonify({'success':True,'logs':[dict(r) for r in rows],'total':total})
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/audit/<int:iid>/export')
def api_audit_export(iid):
    """Export audit log as CSV for compliance reporting."""
    try:
        c    = db()
        rows = c.execute(
            'SELECT id,action,query,mode,result_count,duration_ms,timestamp,ip '
            'FROM audit_log WHERE index_id=? ORDER BY timestamp DESC', (iid,)
        ).fetchall()
        c.close()
        def gen():
            buf = io.StringIO()
            w   = csv.writer(buf)
            w.writerow(['id','action','query','mode','result_count','duration_ms','timestamp','ip'])
            yield buf.getvalue()
            for r in rows:
                buf = io.StringIO(); w = csv.writer(buf)
                w.writerow(list(r)); yield buf.getvalue()
        return Response(gen(), mimetype='text/csv',
                        headers={'Content-Disposition':f'attachment; filename="audit_{iid}.csv"'})
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/preview', methods=['POST'])
def api_preview():
    try:
        data = request.get_json() or {}
        path = data.get('path','')
        if not path or not os.path.exists(path): return jsonify({'error':'文件不存在'}), 404
        st  = os.stat(path)
        ext = os.path.splitext(path)[1].lower()
        info = {
            'path':path, 'name':os.path.basename(path),
            'size_kb':round(st.st_size/1024,2),
            'modified':time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(st.st_mtime)),
            'ext':ext, 'content':None, 'lines':None, 'encoding':None,
        }
        if ext in TEXT_EXTS and st.st_size < 2*1024*1024:
            for enc in ('utf-8','gb18030','latin-1'):
                try:
                    with open(path,'r',encoding=enc,errors='strict') as f:
                        lines = []
                        for i,line in enumerate(f):
                            if i >= 100: break
                            lines.append(line.rstrip('\n'))
                    info['content']  = '\n'.join(lines)
                    info['lines']    = i+1
                    info['encoding'] = enc
                    break
                except Exception: continue
        return jsonify({'success':True,'file':info})
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/duplicates/<int:iid>')
def api_duplicates(iid):
    try:
        c    = db()
        rows = c.execute(
            'SELECT size_hash,COUNT(*) cnt,GROUP_CONCAT(path,"|||") paths,name,size_kb '
            'FROM files WHERE index_id=? GROUP BY size_hash HAVING cnt>1 '
            'ORDER BY size_kb DESC LIMIT 200', (iid,)
        ).fetchall()
        c.close()
        groups = [{'name':r['name'],'size_kb':r['size_kb'],'count':r['cnt'],
                   'paths':r['paths'].split('|||')} for r in rows]
        return jsonify({'success':True,'groups':groups,'total':len(groups)})
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/timeline/<int:iid>')
def api_timeline(iid):
    days = min(90, max(1, int(request.args.get('days',14))))
    try:
        since_ts = int((datetime.now()-timedelta(days=days)).timestamp())
        c    = db()
        rows = c.execute(
            'SELECT name,path,ext,size_kb,modified_time FROM files '
            'WHERE index_id=? AND modified_ts>=? ORDER BY modified_ts DESC LIMIT 200',
            (iid,since_ts)
        ).fetchall()
        c.close()
        return jsonify({'success':True,'files':[dict(r) for r in rows]})
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/export/<int:iid>')
def api_export(iid):
    try:
        c    = db()
        meta = c.execute('SELECT name FROM scan_indexes WHERE id=?',(iid,)).fetchone()
        if not meta: c.close(); return jsonify({'error':'索引不存在'}), 404
        rows = c.execute('SELECT path,name,ext,size_kb,modified_time,creator FROM files WHERE index_id=?',(iid,)).fetchall()
        c.close()
        def gen():
            buf = io.StringIO(); w = csv.writer(buf)
            w.writerow(['path','name','ext','size_kb','modified_time','creator']); yield buf.getvalue()
            for r in rows:
                buf = io.StringIO(); w = csv.writer(buf)
                w.writerow([r['path'],r['name'],r['ext'],r['size_kb'],r['modified_time'],r['creator']])
                yield buf.getvalue()
        safe = secure_filename(meta['name'] or f'index_{iid}')+'.csv'
        return Response(gen(), mimetype='text/csv',
                        headers={'Content-Disposition':f'attachment; filename="{safe}"'})
    except Exception as e: return jsonify({'error':str(e)}), 500


@app.route('/api/settings', methods=['GET'])
def api_get_cfg():
    s = _cfg()
    return jsonify({'success':True,'ignorePatterns':s.get('ignorePatterns',[]),
                    'fuzzyThreshold':s.get('fuzzyThreshold',55),
                    'apiKeyEnabled':bool(s.get('apiKey','').strip()),
                    'fuzzyEngine':FUZZY_ENGINE,'pinyinAvailable':PINYIN_OK,
                    'capabilities':{
                        'pdf':PDF_OK,'docx':DOCX_OK,'xlsx':XLSX_OK,
                    }})

@app.route('/api/settings', methods=['POST'])
def api_save_cfg():
    try:
        data = request.get_json() or {}
        s    = _cfg()
        if 'ignorePatterns' in data: s['ignorePatterns']=[p.strip() for p in data['ignorePatterns'] if p.strip()]
        if 'fuzzyThreshold' in data: s['fuzzyThreshold']=max(0,min(100,int(data['fuzzyThreshold'])))
        if 'apiKey'         in data: s['apiKey']=data['apiKey'].strip()
        _save_cfg(s)
        return jsonify({'success':True})
    except Exception as e: return jsonify({'error':str(e)}), 500

@app.route('/api/status')
def api_status():
    return jsonify({'success':True,'fuzzyEngine':FUZZY_ENGINE,'pinyinAvailable':PINYIN_OK,
                    'capabilities':{'pdf':PDF_OK,'docx':DOCX_OK,'xlsx':XLSX_OK}})

@app.errorhandler(Exception)
def _err(e): return jsonify({'error':'服务器错误','message':str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT',3000))
    print(f'✦ FileIndex Enterprise  →  http://localhost:{port}')
    print(f'  内容提取  PDF:{PDF_OK}  DOCX:{DOCX_OK}  XLSX:{XLSX_OK}')
    print(f'  模糊搜索  {FUZZY_ENGINE}')
    print(f'  拼音支持  {PINYIN_OK}')
    print(f'  数据库    {DB_PATH}')
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
